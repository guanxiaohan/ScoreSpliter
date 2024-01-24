
import openpyxl.worksheet.worksheet
import openpyxl.cell
import openpyxl
import typing
import copy
import re
from collections import OrderedDict
from openpyxl.utils import get_column_letter, column_index_from_string
from time import asctime

class SingleScore(typing.TypedDict):
    SubjectName: str
    Score: int
    ClassRank: int
    SchoolRank: int

class SingleStudent(typing.TypedDict):
    Name: str
    ClassRank: int
    SchoolRank: int
    Scores: dict[SingleScore]

def Log(str_:str = "Logging.") -> None:
    print(f"[{asctime()}] {str_}")

def copy_sheet_contents(sheet:openpyxl.worksheet.worksheet.Worksheet, save_sheet:openpyxl.worksheet.worksheet.Worksheet, source_range=None, tar_range=None):
    sheet_model = sheet
    sheet_new = save_sheet
    sheet_new.sheet_properties.tabColor = sheet_model.sheet_properties.tabColor
    if source_range is not None:
        source_area = sheet_model[source_range]
    else:
        source_area = sheet_model
    merge_cell_dict = OrderedDict()
    merged_ranges = sheet_model.merged_cells.ranges
    for source_row in source_area:
        for source_cell in source_row:
            sc_str = str(source_cell)
            point_time = sc_str.count('.')
            sc_str = sc_str.replace('.', '', point_time - 1)
            start = sc_str.find('.')
            sc_str = sc_str[start + 1: -1]
            for merged_range in merged_ranges:
                if source_cell.coordinate in merged_range:
                    _cell_value = sheet_model.cell(row=merged_range.min_row, column=merged_range.min_col)
                    merge_cell_dict[sc_str] = (merged_range.min_row, merged_range.min_col, _cell_value)
                    continue
    range_li = []
    for val in set(merge_cell_dict.values()):
        tmp = []
        for x, y in merge_cell_dict.items():
            if y == val:
                tmp.append(x)
        if len(tmp):
            range_li.append(min(tmp) + ':' + max(tmp))

    for i in range_li:
        # print(i)
        if source_range is not None:
            base_point_letter = source_range.split(':')[0]
            base_point = sheet_model[base_point_letter]
            base_row = base_point.row
            base_col = base_point.column
        else:
            base_point_letter = i.split(':')[0]
            base_point = sheet_model[base_point_letter]
            base_row = base_point.row
            base_col = base_point.column
        s = i.split(':')[0]
        e = i.split(':')[1]
        # 模板区间第一个点相对顶点距离
        base_delta_row = sheet_model[s].row - base_row
        base_delta_col = sheet_model[s].column - base_col
        # 模板区间两个端点距离
        delta_row = sheet_model[e].row - sheet_model[s].row
        delta_col = sheet_model[e].column - sheet_model[s].column
        # print(base_delta_row, base_delta_col, delta_row, delta_col)
        if tar_range is not None:
            tar_s = tar_range.split(':')[0]
            tar_s_letter = re.findall(r'([A-Za-z]+)', tar_s)[0]
            tar_base_col_idx = column_index_from_string(tar_s_letter)
            tar_base_row_idx = int(re.findall(r'(\d+)', tar_s)[0])
        else:
            tar_s = s
            tar_s_letter = re.findall(r'([A-Za-z]+)', tar_s)[0]
            tar_base_col_idx = column_index_from_string(tar_s_letter)
            tar_base_row_idx = int(re.findall(r'(\d+)', tar_s)[0])
        # print(tar_base_row_idx, tar_base_col_idx)
        tar_range_s_col = get_column_letter(tar_base_col_idx + base_delta_col)
        tar_range_s_idx = tar_base_row_idx + base_delta_row
        tar_range_e_col = get_column_letter(tar_base_col_idx + base_delta_col + delta_col)
        tar_range_e_idx = tar_base_row_idx + base_delta_row + delta_row
        tar_merge = tar_range_s_col + str(tar_range_s_idx) + ':' + tar_range_e_col + str(tar_range_e_idx)
        # print('tar merge:', tar_merge)
        sheet_new.merge_cells(tar_merge)

    if source_range is not None and tar_range is not None:
        source_point_letter = source_range.split(':')[0]
        source_point = sheet_model[source_point_letter]
        source_row = source_point.row
        source_col = source_point.column

        tar_point_letter = tar_range.split(':')[0]
        tar_point = sheet_model[tar_point_letter]
        tar_row = tar_point.row
        tar_col = tar_point.column

        delta_row = tar_row - source_row
        delta_col = tar_col - source_col
        print('ROW:', tar_row, source_row)
        print('COL:', tar_col, source_col)
    else:
        delta_row = 0
        delta_col = 0
    print('DELTA ROW COL:', delta_row, delta_col)

    for source_row in source_area:
        update_row_h = False
        for source_cell in source_row:
            source_x = source_cell.row
            new_x = source_x + delta_row
            source_y = source_cell.column
            new_y = source_y + delta_col

            if not update_row_h:
                sheet_new.row_dimensions[new_x].height = sheet_model.row_dimensions[source_x].height
                update_row_h = True

            sheet_new.column_dimensions[get_column_letter(new_y)].width = \
                sheet_model.column_dimensions[get_column_letter(source_y)].width

            sheet_new.cell(row=new_x, column=new_y, value=source_cell.value)

            # 设置单元格格式
            target_cell = sheet_new.cell(new_x, new_y)
            target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

def parser_merged_cell(sheet: openpyxl.worksheet.worksheet.Worksheet, row, col):
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell

def integerable(value) -> bool:
    try:
        int(value)
        return True
    except:
        return False

def openWorkbook(directory:str) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(directory)
    Log("Opened workbook at: %s"%directory)
    return wb

def detectSheetHead(worksheet:openpyxl.worksheet.worksheet.Worksheet) -> tuple[tuple]:
    dataRegion = (worksheet.max_row, worksheet.max_column)
    startRow = 1
    lastNullRow = 0

    for i in range(1, dataRegion[0]):
        nonintPercent:float = 0
        for j in range(dataRegion[1]):
            allNone = True
            if not integerable(worksheet.cell(i+1, j+1).value):
                nonintPercent += 1/dataRegion[1]
            if worksheet.cell(i+1, j+1).value:
                allNone = False

        if allNone and i-1 == lastNullRow:
            lastNullRow = i

        if nonintPercent < 0.68:
            if i == 2:
                startRow = 1
            elif i - lastNullRow == 1:
                startRow = i-1
            else:
                startRow = lastNullRow+1
            break
        
    return ((startRow, 1), (i, dataRegion[1]))

def detectAvailableSheet(workbook:openpyxl.Workbook) -> str:
    trustables:list[int] = []
    for sheetname in workbook.sheetnames:

        sheet = workbook[sheetname]
        dataRegion = (sheet.max_row, sheet.max_column)
        allCount = dataRegion[0] * dataRegion[1]

        availablePercent:float = 0.0
        similarPercent:float = 0.0

        for i in range(1, dataRegion[0] + 1):
            for j in range(1, dataRegion[1] + 1):
                cell = parser_merged_cell(sheet, i, j)
                if cell.value:
                    availablePercent += 1 / allCount
                else:
                    print(cell)
                
        dataRange = detectSheetHead(sheet)
        allDataCount = 0
        similarDataCount = 0
        for i in range(dataRange[1][0] + 1, dataRegion[0]):
            for j in range(1, dataRegion[1] + 1):
                allDataCount += 1
                cell = parser_merged_cell(sheet, i, j)
                nextcell = parser_merged_cell(sheet, i+1, j)
                if (integerable(cell.value) and integerable(nextcell.value)) or ((not integerable(cell.value)) and (not integerable(nextcell.value))):
                    similarDataCount += 1
                
        similarPercent = similarDataCount/allDataCount if allDataCount > 0 else 0
        print(availablePercent, similarPercent)
        trustables.append(availablePercent + similarPercent)

    return workbook.sheetnames[trustables.index(max(trustables))]

def sortDatas(worksheet_:openpyxl.worksheet.worksheet.Worksheet) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    workbook.create_sheet("Generator Result", 0)
    worksheet = workbook.get_sheet_by_name("Generator Result")
    
    headRange = detectSheetHead(worksheet_)
    headFirstCell = worksheet_.cell(*headRange[0])
    headLastCell = worksheet_.cell(*headRange[1])
    headRows = headRange[1][0]- headRange[0][0] +1
    headRangeStr = f"{headFirstCell.coordinate}:{headLastCell.coordinate}"

    for index, copyFrom in enumerate(range(headRange[1][0] +1, worksheet_.max_row +1)):
        targetHeadFirstCell = worksheet.cell((headRows +1 +2) *index +1, 1)
        targetHeadLastCell = worksheet.cell((headRows +1 +2) *index +headRows, headRange[1][1])
        targetHeadRangeStr = f"{targetHeadFirstCell.coordinate}:{targetHeadLastCell.coordinate}"
        copy_sheet_contents(worksheet_, worksheet, headRangeStr, targetHeadRangeStr)

        originFirstCell = worksheet_.cell(copyFrom, headRange[0][1])
        originLastCell = worksheet_.cell(copyFrom, headRange[1][1])
        originRangeStr = f"{originFirstCell.coordinate}:{originLastCell.coordinate}"
        targetFirstCell = worksheet.cell((headRows +1 +2) *index +headRows +1, 1)
        targetLastCell = worksheet.cell((headRows +1 +2) *index +headRows +1, headRange[1][1])
        targetRangeStr = f"{targetFirstCell.coordinate}:{targetLastCell.coordinate}"
        copy_sheet_contents(worksheet_, worksheet, originRangeStr, targetRangeStr)
        

    return workbook

if __name__ == "__main__":
    ...